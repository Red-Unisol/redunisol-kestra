from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
from collections import Counter
from dataclasses import dataclass, fields
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

DEFAULT_API_BASE_URL = "https://celesol.dyndns.org:5050"
CACHE_SCHEMA_VERSION = 2

SOCIO_FIELDS = [
    "ID",
    "NombreCompleto",
    "NroSocio",
    "NroDoc",
    "CUIT",
]

SOLICITUD_FIELDS = [
    "Oid",
    "NroSolicitud",
    "Fecha",
    "EstadoBase",
    "Estado.Descripcion",
    "CUIT",
    "NroDocumento",
    "NroSocio",
    "MontoAFinanciar",
    "MontoADesembolsar",
    "MontoOriginal",
    "Capital",
    "Prestamo.ID",
    "Prestamo.NroCuenta",
    "Prestamo.FechaEmision",
    "Prestamo.MontoPrestamo",
    "Prestamo.MontoADesembolsar",
    "Prestamo.Referencia",
    "Prestamo.Asiento.NroAsiento",
]

PRESTAMO_FIELDS = [
    "ID",
    "NroCuenta",
    "Referencia",
    "FechaEmision",
    "MontoPrestamo",
    "MontoADesembolsar",
    "Capital",
    "SaldoPrestamo",
    "Estado",
    "SocioTitular.Socio.NombreCompleto",
    "SocioTitular.Socio.NroSocio",
    "SocioTitular.Socio.NroDoc",
    "SocioTitular.Socio.CUIT",
    "Solicitud.Oid",
    "Solicitud.NroSolicitud",
    "Solicitud.Fecha",
    "Solicitud.MontoAFinanciar",
    "Solicitud.MontoADesembolsar",
    "Asiento.NroAsiento",
]

FULL_REPORT_HEADERS = [
    "monto",
    "nroTransaccion",
    "cuitTercero",
    "Nombre",
    "NroSolicitud",
    "NroPrestamo",
    "NroAsiento",
    "FechaMovimiento",
    "ArchivoOrigen",
    "TipoTransaccion",
    "OrigenTransaccion",
    "TitularTercero",
    "FechaBancoReal",
    "ValorFirmado",
    "NroSocio",
    "NroDocumento",
    "SolicitudOid",
    "PrestamoId",
    "FechaSolicitud",
    "FechaPrestamo",
    "MontoSolicitud",
    "MontoADesembolsar",
    "MontoPrestamo",
    "ReferenciaPrestamo",
    "EstadoSolicitud",
    "EstadoPrestamo",
    "MatchEstado",
    "MatchScore",
    "DiasDiferencia",
    "DiferenciaMonto",
    "DiferenciaMontoPct",
    "MontoUsadoParaMatch",
    "FechaUsadaParaMatch",
    "FuenteMatch",
    "ErrorAPI",
]

HIGH_MATCH_REPORT_HEADERS = [
    "monto",
    "nroTransaccion",
    "cuitTercero",
    "Nombre",
    "NroSolicitud",
    "NroPrestamo",
    "NroAsiento",
    "FechaMovimiento",
    "ArchivoOrigen",
    "TipoTransaccion",
    "OrigenTransaccion",
    "TitularTercero",
    "FechaBancoReal",
    "ValorFirmado",
    "NroSocio",
    "NroDocumento",
    "SolicitudOid",
    "PrestamoId",
    "FechaSolicitud",
    "FechaPrestamo",
    "MontoSolicitud",
    "MontoADesembolsar",
    "MontoPrestamo",
    "ReferenciaPrestamo",
    "MatchEstado",
    "MatchScore",
    "MontoUsadoParaMatch",
    "FechaUsadaParaMatch",
    "FuenteMatch",
    "ErrorAPI",
]

MONEY_COLUMNS = {
    "monto",
    "ValorFirmado",
    "MontoSolicitud",
    "MontoADesembolsar",
    "MontoPrestamo",
    "DiferenciaMonto",
    "MontoUsadoParaMatch",
}

DATE_COLUMNS = {
    "FechaMovimiento",
    "FechaSolicitud",
    "FechaPrestamo",
    "FechaUsadaParaMatch",
}

DATETIME_COLUMNS = {"FechaBancoReal"}
PERCENT_COLUMNS = {"DiferenciaMontoPct"}
SCORE_COLUMNS = {"MatchScore"}


def get_report_headers(only_high_matches: bool) -> list[str]:
    return HIGH_MATCH_REPORT_HEADERS if only_high_matches else FULL_REPORT_HEADERS


def filter_report_rows(rows: list[dict[str, Any]], only_high_matches: bool) -> list[dict[str, Any]]:
    if not only_high_matches:
        return rows
    return [row for row in rows if row.get("MatchEstado") == "alto"]


def split_fields(campos: list[str]) -> str:
    return ";".join(campos)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def read_text_with_fallback(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        if "\x00" in text and encoding != "utf-16":
            continue
        return text
    return raw.decode("latin-1", errors="replace")


def normalize_cuit(value: Any) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D+", "", str(value))
    if not digits:
        return None
    return digits


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"*", "null", "None"}:
        return ""
    return text


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = str(value).strip()
    if not text or text in {"*", "null", "None"}:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def normalize_datetime_text(text: str) -> str:
    return re.sub(r"(\.\d{6})\d+", r"\1", text)


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = clean_text(value)
    if not text:
        return None
    text = normalize_datetime_text(text)
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_date(value: Any) -> date | None:
    parsed = parse_datetime(value)
    return parsed.date() if parsed else None


def to_excel_number(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def stringify_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".")
    text = str(value).strip()
    if text in {"*", "null", "None"}:
        return ""
    if re.fullmatch(r"-?\d+(\.0+)?", text):
        return text.split(".")[0]
    return text


def first_decimal(*values: Any) -> Decimal | None:
    for value in values:
        dec = parse_decimal(value)
        if dec is not None:
            return dec
    return None


def first_positive_decimal(*values: Any) -> Decimal | None:
    for value in values:
        dec = parse_decimal(value)
        if dec is not None and dec > 0:
            return dec
    return first_decimal(*values)


def first_non_blank(*values: Any) -> Any:
    for value in values:
        if not is_blank(value):
            return value
    return None


@dataclass(slots=True)
class Movement:
    source_file: str
    fecha_movimiento: date | None
    fecha_banco_real: datetime | None
    tipo_transaccion: str
    origen_transaccion: str
    id_origen_transaccion: str
    monto: Decimal
    valor_firmado: Decimal | None
    nro_transaccion: str
    cuit_tercero: str
    titular_tercero: str


@dataclass(slots=True)
class Candidate:
    source: str
    nombre: str = ""
    nro_socio: str = ""
    nro_documento: str = ""
    cuit: str = ""
    solicitud_oid: str = ""
    nro_solicitud: str = ""
    nro_asiento: str = ""
    fecha_solicitud: date | None = None
    estado_solicitud: str = ""
    prestamo_id: str = ""
    nro_prestamo: str = ""
    referencia_prestamo: str = ""
    fecha_prestamo: date | None = None
    monto_solicitud: Decimal | None = None
    monto_desembolso: Decimal | None = None
    monto_prestamo: Decimal | None = None
    saldo_prestamo: Decimal | None = None
    estado_prestamo: str = ""

    def comparison_date(self) -> date | None:
        return self.fecha_prestamo or self.fecha_solicitud

    def comparison_amount(self) -> Decimal | None:
        return first_positive_decimal(self.monto_desembolso, self.monto_solicitud, self.monto_prestamo)

    def merge_from(self, other: "Candidate") -> None:
        for item in fields(self):
            current = getattr(self, item.name)
            incoming = getattr(other, item.name)
            if is_blank(current) and not is_blank(incoming):
                setattr(self, item.name, incoming)
        if other.source and other.source not in self.source.split("+"):
            self.source = "+".join(sorted(set(self.source.split("+")) | set(other.source.split("+"))))


@dataclass(slots=True)
class MatchEvaluation:
    candidate: Candidate
    score: float
    match_estado: str
    days_diff: int | None
    amount_diff: Decimal | None
    amount_diff_pct: Decimal | None
    comparison_amount: Decimal | None
    comparison_date: date | None


@dataclass(slots=True)
class ApiBundle:
    socio: dict[str, Any] | None
    solicitudes: list[dict[str, Any]]
    prestamos: list[dict[str, Any]]
    errors: list[str]


def map_rows(field_names: list[str], rows: list[Any]) -> list[dict[str, Any]]:
    mapped: list[dict[str, Any]] = []
    for row in rows:
        if isinstance(row, dict):
            mapped.append(row)
            continue
        if not isinstance(row, list):
            raise ValueError(f"Respuesta inesperada de la API: {row!r}")
        mapped.append({field: row[index] if index < len(row) else None for index, field in enumerate(field_names)})
    return mapped


class VimarxClient:
    def __init__(self, base_url: str, timeout: int, max_rows: int, cache_dir: Path | None) -> None:
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self.max_rows = max_rows
        self.cache_dir = cache_dir
        self.session = requests.Session()
        retries = Retry(
            total=4,
            backoff_factor=0.5,
            status_forcelist=(429, 500, 502, 503, 504),
            allowed_methods=frozenset({"POST"}),
        )
        adapter = HTTPAdapter(max_retries=retries)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)
        self.session.verify = False
        if self.cache_dir:
            self.cache_dir.mkdir(parents=True, exist_ok=True)

    def evaluate_list(self, tipo: str, campos: list[str], cmd: str, max_rows: int | None = None) -> list[dict[str, Any]]:
        payload = {
            "cmd": cmd,
            "tipo": tipo,
            "campos": split_fields(campos),
            "max": max_rows if max_rows is not None else self.max_rows,
        }
        response = self.session.post(
            f"{self.base_url}/api/Empresa/EvaluateList",
            json=payload,
            timeout=self.timeout,
        )
        response.raise_for_status()
        data = response.json()
        if not isinstance(data, list):
            raise ValueError(f"Respuesta inesperada para {tipo}: {data!r}")
        return map_rows(campos, data)

    def cache_path(self, cuit: str) -> Path | None:
        if not self.cache_dir:
            return None
        return self.cache_dir / f"{cuit}.json"

    def fetch_bundle(self, cuit: str) -> ApiBundle:
        cache_path = self.cache_path(cuit)
        if cache_path and cache_path.exists():
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            if cached.get("cache_schema_version") == CACHE_SCHEMA_VERSION:
                return ApiBundle(
                    socio=cached.get("socio"),
                    solicitudes=cached.get("solicitudes", []),
                    prestamos=cached.get("prestamos", []),
                    errors=cached.get("errors", []),
                )

        errors: list[str] = []
        socio_rows: list[dict[str, Any]] = []
        solicitudes: list[dict[str, Any]] = []
        prestamos: list[dict[str, Any]] = []

        queries = [
            ("socio", "F.Module.SocioMutual", SOCIO_FIELDS, f"[CUIT]={cuit}", 5),
            ("solicitudes", "PreSolicitud.Module.Solicitud", SOLICITUD_FIELDS, f"[CUIT]={cuit}", self.max_rows),
            (
                "prestamos",
                "F.Module.Cuentas.Prestamos.Prestamo",
                PRESTAMO_FIELDS,
                f"[SocioTitular.Socio.CUIT]={cuit}",
                self.max_rows,
            ),
        ]

        for label, tipo, campos, cmd, max_rows in queries:
            try:
                rows = self.evaluate_list(tipo=tipo, campos=campos, cmd=cmd, max_rows=max_rows)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{label}: {exc}")
                continue
            if label == "socio":
                socio_rows = rows
            elif label == "solicitudes":
                solicitudes = rows
            elif label == "prestamos":
                prestamos = rows

        bundle = ApiBundle(
            socio=socio_rows[0] if socio_rows else None,
            solicitudes=solicitudes,
            prestamos=prestamos,
            errors=errors,
        )
        if cache_path:
            cache_path.write_text(
                json.dumps(
                    {
                        "cache_schema_version": CACHE_SCHEMA_VERSION,
                        "socio": bundle.socio,
                        "solicitudes": bundle.solicitudes,
                        "prestamos": bundle.prestamos,
                        "errors": bundle.errors,
                    },
                    ensure_ascii=False,
                    indent=2,
                    default=str,
                ),
                encoding="utf-8",
            )
        return bundle


def load_movements(input_dir: Path, pattern: str) -> list[Movement]:
    movements: list[Movement] = []
    for path in sorted(input_dir.glob(pattern)):
        reader = csv.DictReader(io.StringIO(read_text_with_fallback(path)), delimiter=";")
        for row in reader:
            cuit = normalize_cuit(row.get("cuitTercero"))
            if not cuit:
                continue
            monto = parse_decimal(row.get("importe"))
            if monto is None:
                continue
            movements.append(
                Movement(
                    source_file=path.name,
                    fecha_movimiento=parse_date(row.get("fechaBanco")),
                    fecha_banco_real=parse_datetime(row.get("fechaBancoReal")),
                    tipo_transaccion=clean_text(row.get("tipoTransaccion")),
                    origen_transaccion=clean_text(row.get("origenTransaccion")),
                    id_origen_transaccion=clean_text(row.get("idOrigenTransaccion")),
                    monto=monto,
                    valor_firmado=parse_decimal(row.get("valor")),
                    nro_transaccion=clean_text(row.get("nroTransaccion")),
                    cuit_tercero=cuit,
                    titular_tercero=clean_text(row.get("titularTercero")),
                )
            )
    return movements


def build_candidates(bundle: ApiBundle) -> list[Candidate]:
    socio = bundle.socio or {}
    candidates: dict[tuple[str, str], Candidate] = {}

    def upsert(key: tuple[str, str], candidate: Candidate) -> None:
        current = candidates.get(key)
        if current:
            current.merge_from(candidate)
        else:
            candidates[key] = candidate

    for row in bundle.solicitudes:
        prestamo_id = stringify_identifier(row.get("Prestamo.ID"))
        key = ("prestamo", prestamo_id) if prestamo_id else ("solicitud", stringify_identifier(row.get("Oid")))
        candidate = Candidate(
            source="solicitud",
            nombre=clean_text(socio.get("NombreCompleto")),
            nro_socio=stringify_identifier(first_non_blank(row.get("NroSocio"), socio.get("NroSocio"))),
            nro_documento=stringify_identifier(first_non_blank(row.get("NroDocumento"), socio.get("NroDoc"))),
            cuit=stringify_identifier(first_non_blank(row.get("CUIT"), socio.get("CUIT"))),
            solicitud_oid=stringify_identifier(row.get("Oid")),
            nro_solicitud=stringify_identifier(row.get("NroSolicitud")),
            nro_asiento=stringify_identifier(row.get("Prestamo.Asiento.NroAsiento")),
            fecha_solicitud=parse_date(row.get("Fecha")),
            estado_solicitud=clean_text(first_non_blank(row.get("Estado.Descripcion"), row.get("EstadoBase"))),
            prestamo_id=prestamo_id,
            nro_prestamo=stringify_identifier(row.get("Prestamo.NroCuenta")),
            referencia_prestamo=stringify_identifier(row.get("Prestamo.Referencia")),
            fecha_prestamo=parse_date(row.get("Prestamo.FechaEmision")),
            monto_solicitud=first_positive_decimal(row.get("MontoAFinanciar"), row.get("MontoOriginal"), row.get("Capital")),
            monto_desembolso=first_positive_decimal(row.get("MontoADesembolsar"), row.get("Prestamo.MontoADesembolsar")),
            monto_prestamo=parse_decimal(row.get("Prestamo.MontoPrestamo")),
        )
        upsert(key, candidate)

    for row in bundle.prestamos:
        prestamo_id = stringify_identifier(row.get("ID"))
        key = ("prestamo", prestamo_id) if prestamo_id else ("prestamo", stringify_identifier(row.get("NroCuenta")))
        candidate = Candidate(
            source="prestamo",
            nombre=clean_text(first_non_blank(row.get("SocioTitular.Socio.NombreCompleto"), socio.get("NombreCompleto"))),
            nro_socio=stringify_identifier(first_non_blank(row.get("SocioTitular.Socio.NroSocio"), socio.get("NroSocio"))),
            nro_documento=stringify_identifier(first_non_blank(row.get("SocioTitular.Socio.NroDoc"), socio.get("NroDoc"))),
            cuit=stringify_identifier(first_non_blank(row.get("SocioTitular.Socio.CUIT"), socio.get("CUIT"))),
            solicitud_oid=stringify_identifier(row.get("Solicitud.Oid")),
            nro_solicitud=stringify_identifier(row.get("Solicitud.NroSolicitud")),
            nro_asiento=stringify_identifier(row.get("Asiento.NroAsiento")),
            fecha_solicitud=parse_date(row.get("Solicitud.Fecha")),
            prestamo_id=prestamo_id,
            nro_prestamo=stringify_identifier(row.get("NroCuenta")),
            referencia_prestamo=stringify_identifier(row.get("Referencia")),
            fecha_prestamo=parse_date(row.get("FechaEmision")),
            monto_solicitud=first_positive_decimal(row.get("Solicitud.MontoAFinanciar")),
            monto_desembolso=first_positive_decimal(row.get("MontoADesembolsar"), row.get("Solicitud.MontoADesembolsar")),
            monto_prestamo=first_positive_decimal(row.get("MontoPrestamo"), row.get("Capital")),
            saldo_prestamo=parse_decimal(row.get("SaldoPrestamo")),
            estado_prestamo=stringify_identifier(row.get("Estado")),
        )
        upsert(key, candidate)

    return list(candidates.values())


def classify_match(days_diff: int | None, amount_diff_pct: Decimal | None, date_window_days: int) -> str:
    if days_diff is None or amount_diff_pct is None:
        return "sin_match_confiable"
    if days_diff <= 7 and amount_diff_pct <= Decimal("0.02"):
        return "alto"
    if days_diff <= min(30, date_window_days) and amount_diff_pct <= Decimal("0.05"):
        return "medio"
    if days_diff <= date_window_days and amount_diff_pct <= Decimal("0.15"):
        return "bajo"
    return "sin_match_confiable"


def evaluate_candidate(movement: Movement, candidate: Candidate, date_window_days: int) -> MatchEvaluation:
    comparison_date = candidate.comparison_date()
    comparison_amount = candidate.comparison_amount()

    days_diff = abs((movement.fecha_movimiento - comparison_date).days) if movement.fecha_movimiento and comparison_date else None

    amount_diff = abs(movement.monto - comparison_amount) if comparison_amount is not None else None
    amount_diff_pct: Decimal | None = None
    if comparison_amount is not None:
        base = max(abs(movement.monto), abs(comparison_amount))
        if base != 0:
            amount_diff_pct = (amount_diff or Decimal("0")) / base

    amount_component = 0.0
    if amount_diff_pct is not None:
        amount_component = max(0.0, 1.0 - min(float(amount_diff_pct), 1.0))

    date_component = 0.0
    if days_diff is not None and date_window_days > 0:
        date_component = max(0.0, 1.0 - min(days_diff, date_window_days * 2) / float(date_window_days * 2))

    score = round((amount_component * 0.7) + (date_component * 0.3), 6)
    match_estado = classify_match(days_diff=days_diff, amount_diff_pct=amount_diff_pct, date_window_days=date_window_days)

    return MatchEvaluation(
        candidate=candidate,
        score=score,
        match_estado=match_estado,
        days_diff=days_diff,
        amount_diff=amount_diff,
        amount_diff_pct=amount_diff_pct,
        comparison_amount=comparison_amount,
        comparison_date=comparison_date,
    )


def select_best_match(movement: Movement, candidates: list[Candidate], date_window_days: int) -> MatchEvaluation | None:
    if not candidates:
        return None
    evaluations = [evaluate_candidate(movement, candidate, date_window_days) for candidate in candidates]
    preferred = [item for item in evaluations if item.match_estado != "sin_match_confiable"]
    pool = preferred if preferred else evaluations
    pool.sort(
        key=lambda item: (
            -item.score,
            item.days_diff if item.days_diff is not None else 10**9,
            float(item.amount_diff_pct) if item.amount_diff_pct is not None else 10**9,
        )
    )
    return pool[0]


def build_report_rows(movements: list[Movement], bundles: dict[str, ApiBundle], date_window_days: int) -> list[dict[str, Any]]:
    report_rows: list[dict[str, Any]] = []

    for movement in sorted(movements, key=lambda item: (item.fecha_movimiento or date.min, item.nro_transaccion)):
        bundle = bundles.get(movement.cuit_tercero, ApiBundle(socio=None, solicitudes=[], prestamos=[], errors=[]))
        candidates = build_candidates(bundle)
        best = select_best_match(movement, candidates, date_window_days)

        if best:
            candidate = best.candidate
            match_estado = best.match_estado
        elif bundle.socio:
            candidate = Candidate(
                source="socio",
                nombre=clean_text(bundle.socio.get("NombreCompleto")),
                nro_socio=stringify_identifier(bundle.socio.get("NroSocio")),
                nro_documento=stringify_identifier(bundle.socio.get("NroDoc")),
                cuit=stringify_identifier(bundle.socio.get("CUIT")),
            )
            match_estado = "sin_creditos"
        else:
            candidate = Candidate(source="")
            match_estado = "sin_socio"

        report_rows.append(
            {
                "monto": to_excel_number(movement.monto),
                "nroTransaccion": movement.nro_transaccion,
                "cuitTercero": movement.cuit_tercero,
                "Nombre": candidate.nombre,
                "NroSolicitud": candidate.nro_solicitud,
                "NroPrestamo": candidate.nro_prestamo,
                "NroAsiento": candidate.nro_asiento,
                "FechaMovimiento": movement.fecha_movimiento,
                "ArchivoOrigen": movement.source_file,
                "TipoTransaccion": movement.tipo_transaccion,
                "OrigenTransaccion": movement.origen_transaccion,
                "TitularTercero": movement.titular_tercero,
                "FechaBancoReal": movement.fecha_banco_real,
                "ValorFirmado": to_excel_number(movement.valor_firmado),
                "NroSocio": candidate.nro_socio,
                "NroDocumento": candidate.nro_documento,
                "SolicitudOid": candidate.solicitud_oid,
                "PrestamoId": candidate.prestamo_id,
                "FechaSolicitud": candidate.fecha_solicitud,
                "FechaPrestamo": candidate.fecha_prestamo,
                "MontoSolicitud": to_excel_number(candidate.monto_solicitud),
                "MontoADesembolsar": to_excel_number(candidate.monto_desembolso),
                "MontoPrestamo": to_excel_number(candidate.monto_prestamo),
                "ReferenciaPrestamo": candidate.referencia_prestamo,
                "EstadoSolicitud": candidate.estado_solicitud,
                "EstadoPrestamo": candidate.estado_prestamo,
                "MatchEstado": match_estado,
                "MatchScore": best.score if best else None,
                "DiasDiferencia": best.days_diff if best else None,
                "DiferenciaMonto": to_excel_number(best.amount_diff if best else None),
                "DiferenciaMontoPct": float(best.amount_diff_pct) if best and best.amount_diff_pct is not None else None,
                "MontoUsadoParaMatch": to_excel_number(best.comparison_amount if best else None),
                "FechaUsadaParaMatch": best.comparison_date if best else None,
                "FuenteMatch": candidate.source,
                "ErrorAPI": " | ".join(bundle.errors),
            }
        )

    return report_rows


def write_excel(output_path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cruce"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=row.get(header))
            if header in MONEY_COLUMNS:
                cell.number_format = '#,##0.00'
            elif header in DATE_COLUMNS:
                cell.number_format = "yyyy-mm-dd"
            elif header in DATETIME_COLUMNS:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif header in PERCENT_COLUMNS:
                cell.number_format = "0.00%"
            elif header in SCORE_COLUMNS:
                cell.number_format = "0.0000"

    for column_index, header in enumerate(headers, start=1):
        max_length = len(header)
        for cell in ws[get_column_letter(column_index)]:
            if cell.value is None:
                continue
            cell_value = cell.value.strftime("%Y-%m-%d %H:%M:%S") if isinstance(cell.value, datetime) else str(cell.value)
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 28)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def print_summary(rows: list[dict[str, Any]], output_path: Path, only_high_matches: bool) -> None:
    counter = Counter(row["MatchEstado"] for row in rows)
    print(f"Excel generado: {output_path}")
    if only_high_matches:
        print("Modo de salida: solo match alto")
    print(f"Filas exportadas: {len(rows)}")
    for key, count in counter.most_common():
        print(f"  {key}: {count}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Cruza los archivos mov_emp con la API Vimarx y exporta un Excel consolidado."
    )
    parser.add_argument("--input-dir", type=Path, default=Path("."), help="Directorio con los archivos mov_emp.")
    parser.add_argument("--pattern", default="mov_emp_*.txt", help="Patrón glob para localizar archivos de entrada.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cruce_mov_emp_vimarx.xlsx"),
        help="Ruta del Excel de salida.",
    )
    parser.add_argument(
        "--api-base-url",
        default=os.getenv("VIMARX_BASE_URL", DEFAULT_API_BASE_URL),
        help="Base URL de la API Evaluate.",
    )
    parser.add_argument("--timeout", type=int, default=30, help="Timeout de cada request a la API, en segundos.")
    parser.add_argument("--max-rows", type=int, default=200, help="Máximo de filas por consulta EvaluateList.")
    parser.add_argument(
        "--date-window-days",
        type=int,
        default=120,
        help="Ventana máxima de días para considerar un match como confiable.",
    )
    parser.add_argument(
        "--cache-dir",
        type=Path,
        default=Path(".cache_vimarx"),
        help="Directorio para cachear respuestas por CUIT.",
    )
    parser.add_argument("--no-cache", action="store_true", help="Desactiva la cache local.")
    parser.add_argument(
        "--only-high-matches",
        "--solo-match-alto",
        dest="only_high_matches",
        action="store_true",
        help="Exporta solo filas con MatchEstado=alto y usa un set reducido de columnas.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    movements = load_movements(args.input_dir, args.pattern)
    if not movements:
        print("No se encontraron movimientos con cuitTercero en los archivos indicados.", file=sys.stderr)
        return 1

    unique_cuits = sorted({movement.cuit_tercero for movement in movements})
    print(f"Movimientos a procesar: {len(movements)}")
    print(f"CUITs únicos a consultar: {len(unique_cuits)}")

    client = VimarxClient(
        base_url=args.api_base_url,
        timeout=args.timeout,
        max_rows=args.max_rows,
        cache_dir=None if args.no_cache else args.cache_dir,
    )

    bundles: dict[str, ApiBundle] = {}
    for index, cuit in enumerate(unique_cuits, start=1):
        bundle = client.fetch_bundle(cuit)
        bundles[cuit] = bundle
        print(
            f"[{index}/{len(unique_cuits)}] {cuit}: "
            f"socio={'si' if bundle.socio else 'no'}, "
            f"solicitudes={len(bundle.solicitudes)}, "
            f"prestamos={len(bundle.prestamos)}, "
            f"errores={len(bundle.errors)}"
        )

    report_rows = build_report_rows(movements, bundles, args.date_window_days)
    report_rows = filter_report_rows(report_rows, args.only_high_matches)
    headers = get_report_headers(args.only_high_matches)
    try:
        write_excel(args.output, report_rows, headers)
    except PermissionError:
        print(
            f"No se pudo escribir {args.output} porque el archivo está en uso. "
            "Cerralo o usá otro nombre con --output.",
            file=sys.stderr,
        )
        return 2
    print_summary(report_rows, args.output, args.only_high_matches)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
