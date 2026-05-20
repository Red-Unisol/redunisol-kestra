#!/usr/bin/env python3
"""Refactored pipeline for generating ARC export workbooks."""
from __future__ import annotations

import calendar as _cal
import datetime as _dt
import json
import logging
import math
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, ROUND_FLOOR, getcontext
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Literal, Optional, Sequence, Tuple

import pandas as pd
import requests
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

# Allow requests to reach the internal HTTPS endpoint without trusted certs.
requests.packages.urllib3.disable_warnings()  # type: ignore[attr-defined]

logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())

getcontext().prec = 28
getcontext().rounding = ROUND_HALF_UP
CENT = Decimal("0.01")
THOUSAND = Decimal("1000")
SHOT_MAX = Decimal("80000.00")
SMALL_SHOT_THRESHOLD = Decimal("500.00")
CLUB_MUTUAL_SHOT_MIN = Decimal("11000.00")
DEFAULT_LINEA_SUPERIOR = "LINEAS CBU BANCOS VARIOS"
CLUB_MUTUAL_LINEA_SUPERIOR = "LINEAS CLUB MUTUAL"
LINEA_SUPERIOR_FIELD = "Prestamo.LineaPrestamo.Superior.Descripcion"

API_URL = "https://celesol.dyndns.org:5050/api/Empresa/EvaluateList"
FILTER_TEMPLATE = (
    "[NroCuota] > 0 AND [SaldoCuota] > 0.0m AND [Fecha] <= #{cutoff_date}# "
    f"AND [{LINEA_SUPERIOR_FIELD}] = '{{linea_superior}}'"
)
FIELDS: Sequence[str] = [
    "Prestamo.SocioTitular.Socio.NroSocio",
    "Prestamo.NroCuenta",
    "Prestamo.Cuotas",
    "Prestamo.LineaPrestamo.Codigo",
    "Prestamo.SocioTitular.Socio.CuentaBancariaHabitual.CBU",
    "Prestamo.SocioTitular.Socio.NroDoc",
    "Prestamo.FechaEmision",
    "SaldoCuota",
    "NroCuota",
    "Fecha",
    "Prestamo.SocioTitular.Socio.CuentaBancariaHabitual.SucursalBanco",
    "Prestamo.SocioTitular.Socio.CuentaBancariaHabitual.NroCuenta",
    "Prestamo.SocioTitular.Socio.SocAux.Caja40",
    "Prestamo.[Cuota Resultante Prestamo]",
    LINEA_SUPERIOR_FIELD,
]
MAX_ROWS = 20_000
COLUMN_ORDER: Sequence[str] = (
    "NROAFI",
    "CONVENIO",
    "NRODEU",
    "CODCONC",
    "FECPROC",
    "VENDEU",
    "NROCUO",
    "CANCUO",
    "IMPORTE",
    "NROSUC",
    "NROCTA",
    "NROPLAN",
    "NROLINEA",
    "NROCBU",
    "CAJA40",
    "FECCOB",
    "NRODOC",
    "MOTIVO",
)
DATE_COLUMNS = ("FECPROC", "VENDEU", "FECCOB")
EXPORT_FILENAMES: Dict[str, str] = {
    "a-enviar": "a-enviar.xlsx",
    "bancor-pero-no-enviamos": "bancor-pero-no-enviamos.xlsx",
    "posiblemente-bancor": "posiblemente-bancor.xlsx",
    "no-bancor": "no-bancor.xlsx",
}
REPORT_FILENAME = "reporte.xlsx"
REPORT_COLUMNS: Sequence[str] = (
    "Número Planilla",
    "Número Afiliado",
    "Número Documento",
    "Línea Código",
    "Clasificación Final",
    "Motivo Clasificación",
    "Clasificación Caja40",
    "Planillas por Socio",
    "¿Es Nuevo o Viejo?",
    "CBU",
    "Clasificación CBU",
    "Caja40",
    "Mes Actual - ¿Se Envía?",
    "Mes Actual - Monto Total",
    "Mes Actual - Disparo Máximo",
    "Mes Actual - Cantidad Disparos",
    "Mes Actual - Detalle Disparos",
    "Mes Pasado - ¿Se Envió?",
    "Mes Pasado - Total Enviado",
    "Mes Pasado - Total Cobrado",
    "Mes Pasado - Total sin Cobrar",
    "Mes Pasado - Máximo Cobrado",
    "Mes Pasado - Cantidad Disparos",
    "Mes Pasado - Respuestas",
    "Mes Pasado - Detalle Disparos",
    "Monto Vencido Mas Última Cuota",
    "Valor Cuota",
    "Cuota y Media",
    "Número Cuota Actual",
    "Tiene R8/Otros?",
)
MOTIVO_LABELS: Dict[str, str] = {
    "caja40_no_enviar": "CAJA40 fuera de rango",
    "sin_intentos_previos": "Sin intentos (planilla vieja)",
    "respuesta_excel_no_permitida": "Respuesta distinta de COB/R10",
    "sin_monto_para_subdividir": "No se pudo generar monto",
    "cbu_especial": "CBU a revisar",
    "cbu_fuera_de_regla": "CBU fuera de Bancor",
}

CATEGORY_LABELS: Dict[str, str] = {
    "a-enviar": "A Enviar",
    "bancor-pero-no-enviamos": "Bancor Pero No Enviamos",
    "posiblemente-bancor": "Posiblemente Bancor",
    "no-bancor": "No Bancor",
}
CLUB_MUTUAL_CATEGORY_LABELS: Dict[str, str] = {
    "a-enviar": "A Enviar",
    "bancor-pero-no-enviamos": "Nacion Pero No Enviamos",
    "posiblemente-bancor": "Posiblemente Nacion",
    "no-bancor": "No Nacion",
}
CBU_CLASSIFICATION_LABELS: Dict[str, str] = {
    "Regular": "Bancor",
    "Fuera de Regla": "No Bancor",
    "Especial": "A revisar",
}
CLUB_MUTUAL_CBU_CLASSIFICATION_LABELS: Dict[str, str] = {
    "Regular": "Nacion",
    "Fuera de Regla": "No Nacion",
    "Especial": "A revisar",
}
DISCARDED_LOG_NAME = "no_bancor_entries_debug.log"

ProgressCallback = Callable[[float, str], None]

_API_ROW_KEYS = (
    "nro_socio",
    "planilla",
    "cant_cuotas",
    "linea_codigo",
    "cbu_transferencia",
    "nro_doc",
    "fecha_emision",
    "saldo_cuota",
    "nro_cuota",
    "fecha_cuota",
    "sucursal_banco",
    "nro_cuenta_banco",
    "caja_40",
    "installment_value",
    "linea_superior",
)

def month_end_for(date_value: _dt.date) -> _dt.date:
    last_day = _cal.monthrange(date_value.year, date_value.month)[1]
    return _dt.date(date_value.year, date_value.month, last_day)


def month_end_for_parts(year: int, month: int) -> _dt.date:
    last_day = _cal.monthrange(year, month)[1]
    return _dt.date(year, month, last_day)


def linea_superior_for_mode(*, club_mutual_mode: bool) -> str:
    if club_mutual_mode:
        return CLUB_MUTUAL_LINEA_SUPERIOR
    return DEFAULT_LINEA_SUPERIOR


def build_filter_for_date(
    cutoff_date: _dt.date,
    *,
    line_superior: str = DEFAULT_LINEA_SUPERIOR,
) -> str:
    escaped_line_superior = line_superior.replace("'", "''")
    return FILTER_TEMPLATE.format(
        cutoff_date=cutoff_date.strftime("%Y-%m-%d"),
        linea_superior=escaped_line_superior,
    )


FILTER = build_filter_for_date(month_end_for(_dt.date.today()))


@dataclass
class Attempt:
    amount: Decimal
    entered: bool
    response: str


@dataclass
class AttemptSummary:
    attempts: List[Attempt] = field(default_factory=list)
    responses: set[str] = field(default_factory=set)

    def register(self, attempt: Attempt) -> None:
        self.attempts.append(attempt)
        if attempt.response:
            self.responses.add(attempt.response)

    @property
    def has_attempts(self) -> bool:
        return bool(self.attempts)

    @property
    def entered_amounts(self) -> List[Decimal]:
        return [attempt.amount for attempt in self.attempts if attempt.entered and attempt.amount > Decimal("0")]

    @property
    def all_amounts(self) -> List[Decimal]:
        return [attempt.amount for attempt in self.attempts if attempt.amount > Decimal("0")]

    @property
    def has_forbidden_responses(self) -> bool:
        if not self.responses:
            return False
        allowed = {"COB", "R10"}
        return any(response not in allowed for response in self.responses)

    def to_loggable(self) -> Dict[str, object]:
        return {
            "attempt_count": len(self.attempts),
            "responses": sorted(self.responses),
            "entered_amounts": [str(value) for value in self.entered_amounts],
        }

    @classmethod
    def empty(cls) -> "AttemptSummary":
        return cls()


@dataclass
class ApiRow:
    nro_socio: str
    planilla: str
    total_installments: Optional[int]
    linea_codigo: str
    cbu: str
    nro_doc: str
    fecha_emision: object
    saldo_cuota: Decimal
    nro_cuota: Optional[int]
    fecha_cuota: object
    sucursal_banco: str
    nro_cuenta_banco: str
    caja40_raw: str
    caja40_int: Optional[int]
    installment_value: Decimal
    linea_superior: str


@dataclass
class ConsolidatedPlanilla:
    planilla: str
    nro_socio: str
    nro_doc: str
    linea_codigo: str
    cbu: str
    fecha_emision: object
    total_installments: Optional[int]
    outstanding_amount: Decimal
    installment_value: Decimal
    caja40_raw: str
    caja40_int: Optional[int]
    sucursal_banco: str
    nro_cuenta_banco: str
    latest_fecha: object
    latest_nro_cuota: Optional[int]
    attempts: AttemptSummary
    loan_status: Literal["new", "old"]
    planillas_per_socio: int = 1


@dataclass
class ClassificationOutcome:
    category: Literal["a-enviar", "bancor-pero-no-enviamos", "posiblemente-bancor", "no-bancor"]
    shots: List[Decimal] = field(default_factory=list)
    reason: Optional[str] = None
    estado_cbu: str = "Regular"
    caja40_clasificacion: str = "GRUPOS 1-20"
    caja40_permitido: bool = True


@dataclass(frozen=True)
class ExportFileResult:
    path: Path
    row_count: int


@dataclass(frozen=True)
class ExportResult:
    files: Dict[str, ExportFileResult]
    discarded_entries: List[Dict[str, object]]
    discarded_log: Optional[Path]
    api_response_path: Path

    @property
    def total_rows(self) -> int:
        return sum(result.row_count for key, result in self.files.items() if key != "reporte")

    @property
    def base_files(self) -> Dict[str, ExportFileResult]:
        return {key: value for key, value in self.files.items() if key != "reporte"}

    @property
    def report_row_count(self) -> int:
        report = self.files.get("reporte")
        return report.row_count if report else 0


def _as_text_or_blank(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def _quantize_currency(value: Decimal) -> Decimal:
    return value.quantize(CENT)


def _round_to_thousand(value: Decimal) -> Decimal:
    if value <= Decimal("0"):
        return Decimal("0")
    units = value / THOUSAND
    rounded = _quantize_currency(units.to_integral_value(rounding="ROUND_CEILING") * THOUSAND)
    return rounded


def _round_down_to_thousand(value: Decimal) -> Decimal:
    if value <= Decimal("0"):
        return Decimal("0")
    units = (value / THOUSAND).to_integral_value(rounding=ROUND_FLOOR)
    return _quantize_currency(units * THOUSAND)


def _normalize_shot_amount(value: Decimal) -> Decimal:
    rounded = _round_to_thousand(value)
    if rounded <= Decimal("0"):
        rounded = _quantize_currency(value)
    if rounded > SHOT_MAX:
        rounded = SHOT_MAX
    return _quantize_currency(rounded)


def _round_base_shot(base_value: Decimal, installment_value: Decimal, max_shots: int) -> Decimal:
    if installment_value <= Decimal("0") or max_shots <= 0:
        return _quantize_currency(base_value)
    rounded = _round_to_thousand(base_value)
    if rounded <= Decimal("0"):
        rounded = _quantize_currency(base_value)
    if rounded > SHOT_MAX:
        rounded = SHOT_MAX
    return _quantize_currency(rounded)


def _max_shots_for_planilla(planilla: ConsolidatedPlanilla) -> int:
    if planilla.planillas_per_socio >= 3:
        return 3
    if planilla.planillas_per_socio == 2:
        return 5
    return 10


def _normalize_club_mutual_shot_amount(
    value: Decimal,
    *,
    amount_to_collect: Decimal,
) -> Decimal:
    rounded = _round_to_thousand(value)
    if rounded <= Decimal("0"):
        rounded = _quantize_currency(value)
    if amount_to_collect >= CLUB_MUTUAL_SHOT_MIN and rounded < CLUB_MUTUAL_SHOT_MIN:
        rounded = CLUB_MUTUAL_SHOT_MIN
    return _quantize_currency(rounded)

def _to_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return _quantize_currency(value)
    if isinstance(value, (int, str)):
        try:
            return _quantize_currency(Decimal(str(value)))
        except InvalidOperation:
            return Decimal("0")
    if isinstance(value, float):
        if math.isnan(value):
            return Decimal("0")
        return _quantize_currency(Decimal(str(value)))
    return Decimal("0")


def _normalize_planilla(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, Decimal)):
        return str(int(value))
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(int(round(value)))
    return str(value).strip()


def _parse_optional_int(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    try:
        text = str(value).strip()
    except Exception:
        return None
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _normalize_response(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text.upper()


def _normalize_cbu(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_timestamp(value: object) -> pd.Timestamp:
    if value is None or value == "":
        return pd.Timestamp.min
    try:
        ts = pd.to_datetime(value, errors="coerce")
    except Exception:
        return pd.Timestamp.min
    if pd.isna(ts):
        return pd.Timestamp.min
    return ts


def _is_caja40_allowed(value: Optional[int], *, arrastre_mode: bool) -> bool:
    if arrastre_mode:
        return value in {None, 0, 50, 60, 90}
    return value is not None and 1 <= value <= 20


def _describe_cbu_status(raw_cbu: str, *, club_mutual_mode: bool = False) -> str:
    if not raw_cbu:
        return "Especial"
    cbu_upper = raw_cbu.upper()
    if raw_cbu.startswith("000") or cbu_upper == "REVISAR CBU":
        return "Especial"
    valid_prefix = "011" if club_mutual_mode else "0200"
    if raw_cbu.startswith(valid_prefix):
        return "Regular"
    return "Fuera de Regla"


def _clasificar_caja40(value: Optional[int]) -> str:
    if value is None:
        return "ARRASTRE"
    if 1 <= value <= 20:
        return "GRUPOS 1-20"
    if value in {0, 50, 60, 90}:
        return "ARRASTRE"
    return "CAJAS SIN USO"


def read_excel_attempts(path: Path) -> Dict[str, AttemptSummary]:
    if not path.exists():
        raise FileNotFoundError(f"Excel input not found: {path}")

    df = pd.read_excel(path)
    column_lookup = {str(column).strip().lower(): column for column in df.columns}
    required = {"planilla", "respuesta", "importe"}
    if not required.issubset(column_lookup):
        missing = required.difference(column_lookup)
        raise ValueError(f"Excel file missing required columns: {', '.join(sorted(missing))}")

    planilla_col = column_lookup["planilla"]
    respuesta_col = column_lookup["respuesta"]
    importe_col = column_lookup["importe"]

    grouped: Dict[str, AttemptSummary] = defaultdict(AttemptSummary)

    for planilla_raw, importe_raw, respuesta_raw in zip(
        df[planilla_col], df[importe_col], df[respuesta_col], strict=False
    ):
        planilla = _normalize_planilla(planilla_raw)
        if not planilla:
            continue
        response = _normalize_response(respuesta_raw)
        amount = _to_decimal(importe_raw)
        attempt = Attempt(amount=amount, entered=response == "COB", response=response)
        grouped[planilla].register(attempt)

    return grouped


def fetch_rows(
    filter_expr: str = FILTER,
    fields: Sequence[str] = FIELDS,
    max_rows: int = MAX_ROWS,
    *,
    timeout: int = 300,
    return_raw_text: bool = False,
) -> List[Sequence[object]] | Tuple[List[Sequence[object]], str]:
    payload = {
        "cmd": filter_expr,
        "tipo": "F.Module.Cuentas.Prestamos.CuotaPrestamo",
        "campos": ";".join(fields),
        "max": max_rows,
    }
    logger.debug("Posting EvaluateList payload with max=%s", max_rows)
    response = requests.post(API_URL, json=payload, timeout=timeout, verify=False)
    response.raise_for_status()
    raw_text = response.text
    data = response.json()
    if not isinstance(data, list):
        raise RuntimeError("Unexpected response payload (expected list)")
    if return_raw_text:
        return data, raw_text  # type: ignore[return-value]
    return data  # type: ignore[return-value]



def load_rows_from_dump(path: Path) -> tuple[list[Sequence[object]], str]:
    """Load EvaluateList rows from a cached dump file."""
    dump_path = Path(path)
    if not dump_path.exists():
        raise FileNotFoundError(f"API dump not found: {dump_path}")
    raw_text = dump_path.read_text(encoding="utf-8")
    lines = [
        line
        for line in raw_text.splitlines()
        if not line.lstrip().startswith("//") and line.strip()
    ]
    json_text = "\n".join(lines)
    try:
        data = json.loads(json_text)
    except json.JSONDecodeError as err:
        raise ValueError(f"Invalid JSON content in {dump_path}: {err}") from err
    if not isinstance(data, list):
        raise ValueError(f"Dump file {dump_path} does not contain a JSON list.")
    return data, raw_text


def parse_api_rows(rows: Iterable[Sequence[object]]) -> List[ApiRow]:
    parsed: List[ApiRow] = []
    for row in rows:
        row_map = dict(zip(_API_ROW_KEYS, row, strict=False))
        planilla = _normalize_planilla(row_map.get("planilla"))
        if not planilla:
            logger.debug("Discarding API row without planilla: %s", row_map)
            continue
        parsed.append(
            ApiRow(
                nro_socio=_as_text_or_blank(row_map.get("nro_socio")),
                planilla=planilla,
                total_installments=_parse_optional_int(row_map.get("cant_cuotas")),
                linea_codigo=_as_text_or_blank(row_map.get("linea_codigo")),
                cbu=_normalize_cbu(row_map.get("cbu_transferencia")),
                nro_doc=_as_text_or_blank(row_map.get("nro_doc")),
                fecha_emision=row_map.get("fecha_emision"),
                saldo_cuota=_to_decimal(row_map.get("saldo_cuota")),
                nro_cuota=_parse_optional_int(row_map.get("nro_cuota")),
                fecha_cuota=row_map.get("fecha_cuota"),
                sucursal_banco=_as_text_or_blank(row_map.get("sucursal_banco")),
                nro_cuenta_banco=_as_text_or_blank(row_map.get("nro_cuenta_banco")),
                caja40_raw=_as_text_or_blank(row_map.get("caja_40")),
                caja40_int=_parse_optional_int(row_map.get("caja_40")),
                installment_value=_to_decimal(row_map.get("installment_value")),
                linea_superior=_as_text_or_blank(row_map.get("linea_superior")).strip(),
            )
        )
    logger.debug("Parsed %s API rows", len(parsed))
    return parsed


def filter_api_rows_for_mode(
    api_rows: Sequence[ApiRow],
    *,
    club_mutual_mode: bool,
) -> List[ApiRow]:
    if not club_mutual_mode:
        return list(api_rows)

    filtered = [row for row in api_rows if row.linea_superior.strip() == CLUB_MUTUAL_LINEA_SUPERIOR]
    if filtered:
        return filtered

    if api_rows and not any(row.linea_superior.strip() for row in api_rows):
        raise RuntimeError(
            "El modo Club Mutual requiere que la respuesta incluya "
            f"{LINEA_SUPERIOR_FIELD} en la proyeccion del API o dump."
        )
    return filtered


def sort_api_rows(rows: Iterable[Sequence[object]]) -> List[Sequence[object]]:
    """Return a deterministically ordered copy of the API rows."""

    def sort_key(row: Sequence[object]) -> tuple:
        mapping = dict(zip(_API_ROW_KEYS, row, strict=False))
        planilla = _normalize_planilla(mapping.get("planilla"))
        nro_cuota = _parse_optional_int(mapping.get("nro_cuota")) or 0
        fecha_cuota = _coerce_timestamp(mapping.get("fecha_cuota"))
        saldo = _to_decimal(mapping.get("saldo_cuota"))
        return planilla, nro_cuota, fecha_cuota, saldo

    sorted_rows = sorted(rows, key=sort_key)
    return [list(row) if not isinstance(row, list) else row for row in sorted_rows]


def consolidate_planillas(
    api_rows: Sequence[ApiRow],
    attempts_by_planilla: Dict[str, AttemptSummary],
) -> List[ConsolidatedPlanilla]:
    grouped_rows: Dict[str, List[ApiRow]] = defaultdict(list)
    for row in api_rows:
        grouped_rows[row.planilla].append(row)

    consolidated: List[ConsolidatedPlanilla] = []
    for planilla, rows in grouped_rows.items():
        outstanding = _quantize_currency(sum((row.saldo_cuota for row in rows), Decimal("0")))
        newest = max(rows, key=lambda item: (_coerce_timestamp(item.fecha_cuota), item.nro_cuota or -1))
        attempts = attempts_by_planilla.get(planilla, AttemptSummary.empty())
        loan_status: Literal["new", "old"] = (
            "new" if len(rows) == 1 and (rows[0].nro_cuota == 1) else "old"
        )
        consolidated.append(
            ConsolidatedPlanilla(
                planilla=planilla,
                nro_socio=newest.nro_socio,
                nro_doc=newest.nro_doc,
                linea_codigo=newest.linea_codigo,
                cbu=newest.cbu,
                fecha_emision=newest.fecha_emision,
                total_installments=newest.total_installments,
                outstanding_amount=outstanding,
                installment_value=newest.installment_value or rows[0].installment_value,
                caja40_raw=newest.caja40_raw,
                caja40_int=newest.caja40_int,
                sucursal_banco=newest.sucursal_banco,
                nro_cuenta_banco=newest.nro_cuenta_banco,
                latest_fecha=newest.fecha_cuota,
                latest_nro_cuota=newest.nro_cuota,
                attempts=attempts,
                loan_status=loan_status,
            )
        )
    socio_counts = Counter(cp.nro_socio for cp in consolidated if cp.nro_socio)
    for item in consolidated:
        item.planillas_per_socio = socio_counts.get(item.nro_socio, 1)
    logger.debug("Consolidated into %s planillas", len(consolidated))
    return consolidated


def _compute_standard_shots(planilla: ConsolidatedPlanilla) -> List[Decimal]:
    amount_cap = _quantize_currency(planilla.installment_value * Decimal("1.5"))
    amount_to_collect = min(planilla.outstanding_amount, amount_cap)
    amount_to_collect = _quantize_currency(amount_to_collect)
    if amount_to_collect <= Decimal("0"):
        return []

    if planilla.loan_status == "new":
        shot_amount = Decimal("80000.00")
    else:
        if not planilla.attempts.has_attempts:
            return []
        entered = planilla.attempts.entered_amounts
        if entered:
            shot_amount = max(entered)
        else:
            attempts = planilla.attempts.all_amounts
            if not attempts:
                return []
            shot_amount = attempts[0]
            for amount in attempts[1:]:
                if amount < shot_amount:
                    shot_amount = amount
            shot_amount = shot_amount / Decimal("2")
    shot_amount = _normalize_shot_amount(shot_amount)
    if shot_amount <= Decimal("0"):
        return []

    max_shots = _max_shots_for_planilla(planilla)

    n_full = int(amount_to_collect // shot_amount)
    remainder = _quantize_currency(amount_to_collect - (shot_amount * n_full))
    shots: List[Decimal] = []
    if n_full:
        shots.extend([shot_amount] * n_full)
    if remainder > Decimal("0"):
        shots.append(remainder)
    if not shots:
        shots = [amount_to_collect]

    if len(shots) > max_shots:
        if planilla.installment_value > Decimal("0") and amount_to_collect >= planilla.installment_value:
            cap_total = _quantize_currency(shot_amount * Decimal(max_shots))
            if cap_total >= planilla.installment_value:
                shots = [shot_amount] * max_shots
            else:
                base_value = planilla.installment_value / Decimal(max_shots)
                base_shot = _round_base_shot(base_value, planilla.installment_value, max_shots)
                shots = [base_shot] * (max_shots - 1)
                remainder_value = planilla.installment_value - sum(shots, Decimal("0"))
                if remainder_value > Decimal("0"):
                    shots.append(_quantize_currency(remainder_value))
                else:
                    adjusted_last = _quantize_currency(base_shot + remainder_value)
                    if adjusted_last <= Decimal("0"):
                        adjusted_last = base_shot
                    shots[-1] = adjusted_last
            amount_to_collect = _quantize_currency(sum(shots, Decimal("0")))
        else:
            keep = max_shots - 1
            trimmed = shots[:keep] if keep > 0 else []
            tail_sum = _quantize_currency(sum(shots[keep:], Decimal("0")))
            if keep > 0:
                trimmed.append(tail_sum)
                shots = trimmed
            else:
                shots = [tail_sum]

    total_collected = _quantize_currency(sum(shots, Decimal("0")))
    if (
        planilla.installment_value > Decimal("0")
        and amount_to_collect >= planilla.installment_value
        and total_collected < planilla.installment_value
    ):
        k = len(shots)
        if k <= 1:
            shots = [amount_to_collect]
        else:
            candidate = _normalize_shot_amount(_quantize_currency(planilla.installment_value / Decimal(k)))
            target_shot = max(shot_amount, candidate)
            adjusted = [target_shot] * (k - 1)
            subtotal = _quantize_currency(sum(adjusted, Decimal("0")))
            if subtotal >= amount_to_collect:
                shots = [amount_to_collect]
            else:
                last_value = _quantize_currency(amount_to_collect - subtotal)
                if last_value <= Decimal("0"):
                    shots = [amount_to_collect]
                else:
                    adjusted.append(last_value)
                    shots = adjusted

    # Handle very small remainder shots
    if shots:
        last_value = shots[-1]
        if last_value < SMALL_SHOT_THRESHOLD:
            first_value = shots[0]
            if first_value + last_value <= SHOT_MAX:
                shots[0] = _quantize_currency(first_value + last_value)
                shots.pop()
            elif len(shots) >= 2:
                prev_idx = len(shots) - 2
                target_last = max(SMALL_SHOT_THRESHOLD, _round_to_thousand(last_value))
                if target_last > shots[prev_idx]:
                    shots[prev_idx] = _quantize_currency(shots[prev_idx] + last_value)
                    shots.pop()
                else:
                    needed = target_last - last_value
                    if needed > Decimal("0"):
                        shots[prev_idx] = _quantize_currency(shots[prev_idx] - needed)
                        shots[-1] = _quantize_currency(last_value + needed)

    correction = _quantize_currency(amount_to_collect - sum(shots, Decimal("0")))
    if shots:
        shots[-1] = _quantize_currency(shots[-1] + correction)
    return shots


def _compute_club_mutual_shots(planilla: ConsolidatedPlanilla) -> List[Decimal]:
    amount_cap = _quantize_currency(planilla.installment_value * Decimal("1.5"))
    amount_to_collect = min(planilla.outstanding_amount, amount_cap)
    amount_to_collect = _quantize_currency(amount_to_collect)
    if amount_to_collect <= Decimal("0"):
        return []

    if planilla.loan_status == "new":
        shot_amount = amount_to_collect
    else:
        if not planilla.attempts.has_attempts:
            return []
        entered = planilla.attempts.entered_amounts
        if entered:
            shot_amount = max(entered)
        else:
            attempts = planilla.attempts.all_amounts
            if not attempts:
                return []
            shot_amount = min(attempts) / Decimal("2")

    shot_amount = _normalize_club_mutual_shot_amount(
        shot_amount,
        amount_to_collect=amount_to_collect,
    )
    if shot_amount <= Decimal("0"):
        return []

    max_shots = _max_shots_for_planilla(planilla)
    if amount_to_collect >= CLUB_MUTUAL_SHOT_MIN:
        max_shots = min(max_shots, max(1, int(amount_to_collect // CLUB_MUTUAL_SHOT_MIN)))
    max_shots = max(1, max_shots)

    n_full = int(amount_to_collect // shot_amount)
    remainder = _quantize_currency(amount_to_collect - (shot_amount * n_full))
    shots: List[Decimal] = []
    if n_full:
        shots.extend([shot_amount] * n_full)
    if remainder > Decimal("0"):
        shots.append(remainder)
    if not shots:
        shots = [amount_to_collect]

    if len(shots) > max_shots:
        keep = max_shots - 1
        trimmed = shots[:keep] if keep > 0 else []
        tail_sum = _quantize_currency(sum(shots[keep:], Decimal("0")))
        if keep > 0:
            trimmed.append(tail_sum)
            shots = trimmed
        else:
            shots = [tail_sum]

    if amount_to_collect >= CLUB_MUTUAL_SHOT_MIN and len(shots) > 1 and shots[-1] < CLUB_MUTUAL_SHOT_MIN:
        shots[0] = _quantize_currency(shots[0] + shots[-1])
        shots.pop()

    correction = _quantize_currency(amount_to_collect - sum(shots, Decimal("0")))
    if shots:
        shots[-1] = _quantize_currency(shots[-1] + correction)
    return shots


def compute_shots(
    planilla: ConsolidatedPlanilla,
    *,
    club_mutual_mode: bool = False,
) -> List[Decimal]:
    if club_mutual_mode:
        return _compute_club_mutual_shots(planilla)
    return _compute_standard_shots(planilla)


def determine_planilla_outcome(
    planilla: ConsolidatedPlanilla,
    *,
    arrastre_mode: bool,
    club_mutual_mode: bool = False,
) -> ClassificationOutcome:
    cbu_clean = planilla.cbu.strip() if planilla.cbu else ""
    estado_cbu = _describe_cbu_status(cbu_clean, club_mutual_mode=club_mutual_mode)
    caja40_permitido = _is_caja40_allowed(planilla.caja40_int, arrastre_mode=arrastre_mode)
    caja40_clasificacion = _clasificar_caja40(planilla.caja40_int)

    if estado_cbu == "Especial":
        return ClassificationOutcome(
            category="posiblemente-bancor",
            reason="cbu_especial",
            estado_cbu=estado_cbu,
            caja40_clasificacion=caja40_clasificacion,
            caja40_permitido=caja40_permitido,
        )

    if estado_cbu == "Fuera de Regla":
        return ClassificationOutcome(
            category="no-bancor",
            reason="cbu_fuera_de_regla",
            estado_cbu=estado_cbu,
            caja40_clasificacion=caja40_clasificacion,
            caja40_permitido=caja40_permitido,
        )

    if not caja40_permitido:
        return ClassificationOutcome(
            category="bancor-pero-no-enviamos",
            reason="caja40_no_enviar",
            estado_cbu=estado_cbu,
            caja40_clasificacion=caja40_clasificacion,
            caja40_permitido=caja40_permitido,
        )

    attempts = planilla.attempts
    if planilla.loan_status == "old" and not attempts.has_attempts:
        return ClassificationOutcome(
            category="bancor-pero-no-enviamos",
            reason="sin_intentos_previos",
            estado_cbu=estado_cbu,
            caja40_clasificacion=caja40_clasificacion,
            caja40_permitido=caja40_permitido,
        )
    if attempts.has_forbidden_responses:
        return ClassificationOutcome(
            category="bancor-pero-no-enviamos",
            reason="respuesta_excel_no_permitida",
            estado_cbu=estado_cbu,
            caja40_clasificacion=caja40_clasificacion,
            caja40_permitido=caja40_permitido,
        )

    shots = compute_shots(planilla, club_mutual_mode=club_mutual_mode)
    if not shots:
        return ClassificationOutcome(
            category="bancor-pero-no-enviamos",
            reason="sin_monto_para_subdividir",
            estado_cbu=estado_cbu,
            caja40_clasificacion=caja40_clasificacion,
            caja40_permitido=caja40_permitido,
        )

    return ClassificationOutcome(
        category="a-enviar",
        shots=shots,
        estado_cbu=estado_cbu,
        caja40_clasificacion=caja40_clasificacion,
        caja40_permitido=caja40_permitido,
    )


def _build_export_record(
    planilla: ConsolidatedPlanilla,
    importe: Decimal,
    report_date: _dt.date,
    *,
    reason: str = "",
    ven_date_override: _dt.date | None = None,
) -> Dict[str, object]:
    nro_plan = planilla.planilla
    nro_cuota = planilla.latest_nro_cuota
    if nro_plan and nro_cuota is not None:
        nro_deu = f"{nro_plan}{nro_cuota:03d}"
    else:
        nro_deu = ""
    record = {
        "NROAFI": planilla.nro_socio,
        "CONVENIO": 2,
        "NRODEU": nro_deu,
        "CODCONC": 30,
        "FECPROC": planilla.fecha_emision,
        "VENDEU": ven_date_override or planilla.latest_fecha,
        "NROCUO": nro_cuota or "",
        "CANCUO": planilla.total_installments or "",
        "IMPORTE": _quantize_currency(importe),
        "NROSUC": _as_text_or_blank(planilla.sucursal_banco),
        "NROCTA": _as_text_or_blank(planilla.nro_cuenta_banco),
        "NROPLAN": nro_plan,
        "NROLINEA": planilla.linea_codigo,
        "NROCBU": planilla.cbu,
        "CAJA40": _as_text_or_blank(planilla.caja40_raw),
        "FECCOB": report_date,
        "NRODOC": planilla.nro_doc,
        "MOTIVO": reason or "",
    }
    return record


def _build_report_record(
    planilla: ConsolidatedPlanilla,
    outcome: ClassificationOutcome,
    report_date: _dt.date,
    *,
    club_mutual_mode: bool = False,
) -> Dict[str, object]:
    shots = [_quantize_currency(shot) for shot in outcome.shots]
    monto_total_enviado = _quantize_currency(sum(shots, Decimal("0"))) if shots else Decimal("0.00")
    disparo_maximo_mes_actual = _quantize_currency(max(shots)) if shots else Decimal("0.00")
    se_envia_mes_actual = "SI" if monto_total_enviado > Decimal("0") else "NO"

    excel_amounts = [_quantize_currency(amount) for amount in planilla.attempts.all_amounts]
    importe_mes_anterior = _quantize_currency(sum(excel_amounts, Decimal("0"))) if excel_amounts else Decimal("0.00")
    se_envio_mes_anterior = "SI" if importe_mes_anterior > Decimal("0") else "NO"
    entered_amounts = [
        _quantize_currency(attempt.amount)
        for attempt in planilla.attempts.attempts
        if attempt.entered and attempt.amount > Decimal("0")
    ]
    not_entered_amounts = [
        _quantize_currency(attempt.amount)
        for attempt in planilla.attempts.attempts
        if not attempt.entered and attempt.amount > Decimal("0")
    ]
    total_cobrado = _quantize_currency(sum(entered_amounts, Decimal("0")))
    total_no_cobrado = _quantize_currency(sum(not_entered_amounts, Decimal("0")))

    detalle_disparos_actual = ", ".join(format(shot, '.2f') for shot in shots) if shots else ""

    detalle_pasado_parts: List[str] = []
    for attempt in planilla.attempts.attempts:
        amount = _quantize_currency(attempt.amount)
        if amount <= Decimal("0"):
            continue
        part = format(amount, '.2f')
        response = attempt.response.strip()
        if response:
            part = f"{part} ({response})"
        detalle_pasado_parts.append(part)
    detalle_disparos_pasado = ", ".join(detalle_pasado_parts)

    cuota_y_media = _quantize_currency(planilla.installment_value * Decimal("1.5"))
    intentos_total = len(planilla.attempts.attempts)
    respuestas = ", ".join(sorted(planilla.attempts.responses)) if planilla.attempts.responses else ""
    ultimo_importe_cobrado = _quantize_currency(max(entered_amounts)) if entered_amounts else Decimal("0.00")
    tiene_r8 = "SI" if planilla.attempts.has_forbidden_responses else "NO"
    category_labels = CLUB_MUTUAL_CATEGORY_LABELS if club_mutual_mode else CATEGORY_LABELS
    clasificacion = category_labels.get(outcome.category, outcome.category)
    estado_prestamo = "Nuevo" if planilla.loan_status == "new" else "Viejo"
    cbu_labels = CLUB_MUTUAL_CBU_CLASSIFICATION_LABELS if club_mutual_mode else CBU_CLASSIFICATION_LABELS
    estado_cbu = cbu_labels.get(outcome.estado_cbu, outcome.estado_cbu)
    motivo_labels = MOTIVO_LABELS
    if club_mutual_mode and outcome.reason == "cbu_fuera_de_regla":
        motivo_clasificacion = "CBU fuera de Nacion"
    else:
        motivo_clasificacion = motivo_labels.get(outcome.reason or "", "Incluido en Criterio de Envío")

    return {
        "Número Planilla": planilla.planilla,
        "Número Afiliado": planilla.nro_socio,
        "Número Documento": planilla.nro_doc,
        "Línea Código": planilla.linea_codigo,
        "Clasificación Final": clasificacion,
        "Motivo Clasificación": motivo_clasificacion,
        "Clasificación Caja40": outcome.caja40_clasificacion,
        "Planillas por Socio": planilla.planillas_per_socio,
        "¿Es Nuevo o Viejo?": estado_prestamo,
        "CBU": planilla.cbu,
        "Clasificación CBU": estado_cbu,
        "Caja40": planilla.caja40_raw,
        "Mes Actual - ¿Se Envía?": se_envia_mes_actual,
        "Mes Actual - Monto Total": monto_total_enviado,
        "Mes Actual - Disparo Máximo": disparo_maximo_mes_actual,
        "Mes Pasado - ¿Se Envió?": se_envio_mes_anterior,
        "Mes Pasado - Total Enviado": importe_mes_anterior,
        "Mes Actual - Cantidad Disparos": len(shots),
        "Mes Actual - Detalle Disparos": detalle_disparos_actual,
        "Monto Vencido Mas Última Cuota": _quantize_currency(planilla.outstanding_amount),
        "Valor Cuota": _quantize_currency(planilla.installment_value),
        "Cuota y Media": cuota_y_media,
        "Número Cuota Actual": planilla.latest_nro_cuota or "",
        "Mes Pasado - Cantidad Disparos": intentos_total,
        "Mes Pasado - Respuestas": respuestas,
        "Mes Pasado - Detalle Disparos": detalle_disparos_pasado,
        "Mes Pasado - Total Cobrado": total_cobrado,
        "Mes Pasado - Total sin Cobrar": total_no_cobrado,
        "Mes Pasado - Máximo Cobrado": ultimo_importe_cobrado,
        "Tiene R8/Otros?": tiene_r8,
    }


def classify_planillas(
    planillas: Sequence[ConsolidatedPlanilla],
    report_date: _dt.date,
    *,
    dev_mode: bool = False,
    arrastre_mode: bool = False,
    club_mutual_mode: bool = False,
) -> Tuple[Dict[str, List[Dict[str, object]]], List[Dict[str, object]], List[Dict[str, object]]]:
    outputs: Dict[str, List[Dict[str, object]]] = {key: [] for key in EXPORT_FILENAMES}
    discarded_entries: List[Dict[str, object]] = []
    report_rows: List[Dict[str, object]] = []

    for planilla in planillas:
        outcome = determine_planilla_outcome(
            planilla,
            arrastre_mode=arrastre_mode,
            club_mutual_mode=club_mutual_mode,
        )
        report_rows.append(
            _build_report_record(
                planilla,
                outcome,
                report_date,
                club_mutual_mode=club_mutual_mode,
            )
        )
        if outcome.category == "no-bancor":
            if dev_mode:
                discarded_entries.append(
                    {
                        "planilla": planilla.planilla,
                        "reason": outcome.reason,
                        "cbu": planilla.cbu,
                        "loan_status": planilla.loan_status,
                        "outstanding_amount": str(planilla.outstanding_amount),
                        "installment_value": str(planilla.installment_value),
                        "attempts": planilla.attempts.to_loggable(),
                    }
                )
            outputs[outcome.category].append(
                _build_export_record(
                    planilla,
                    planilla.outstanding_amount,
                    report_date,
                    reason=outcome.reason or "",
                )
            )
            logger.debug("Planilla %s clasificada como descartada (%s)", planilla.planilla, outcome.reason)
            continue

        if outcome.category in ("bancor-pero-no-enviamos", "posiblemente-bancor"):
            outputs[outcome.category].append(
                _build_export_record(
                    planilla,
                    planilla.outstanding_amount,
                    report_date,
                    reason=outcome.reason or "",
                )
            )
            logger.debug(
                "Planilla %s clasificada como %s (%s)",
                planilla.planilla,
                outcome.category,
                outcome.reason,
            )
            continue

        if outcome.category == "a-enviar":
            for idx, shot in enumerate(outcome.shots):
                ven_date = report_date - _dt.timedelta(days=30 * idx)
                outputs[outcome.category].append(_build_export_record(planilla, shot, report_date, ven_date_override=ven_date))
        else:
            for shot in outcome.shots:
                outputs[outcome.category].append(_build_export_record(planilla, shot, report_date))
        logger.debug(
            "Planilla %s -> %s con %s disparos",
            planilla.planilla,
            outcome.category,
            len(outcome.shots),
        )

    return outputs, discarded_entries, report_rows


def export_to_excel(df: pd.DataFrame, path: Path) -> None:
    frame = df.copy()
    if "IMPORTE" in frame.columns:
        frame["IMPORTE"] = frame["IMPORTE"].apply(lambda v: float(_to_decimal(v)))
    for col in DATE_COLUMNS:
        if col in frame.columns:
            frame[col] = pd.to_datetime(frame[col], errors="coerce")

    row_count = len(frame)
    with pd.ExcelWriter(path, engine="openpyxl", date_format="yyyy-mm-dd") as writer:
        frame.to_excel(writer, sheet_name="Sheet1", index=False)
        worksheet = writer.sheets["Sheet1"]
        for column_letter in ("E", "F", "P"):
            for row_idx in range(2, row_count + 2):
                worksheet[f"{column_letter}{row_idx}"].number_format = "yyyy-mm-dd"


def next_available_path(base_path: Path) -> Path:
    if not base_path.exists():
        return base_path
    suffix = 1
    while True:
        candidate = base_path.with_stem(f"{base_path.stem}_{suffix}")
        if not candidate.exists():
            return candidate
        suffix += 1


def write_output_files(
    outputs: Dict[str, List[Dict[str, object]]],
    output_dir: Path,
) -> Dict[str, ExportFileResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: Dict[str, ExportFileResult] = {}
    for category, records in outputs.items():
        df = pd.DataFrame(records) if records else pd.DataFrame(columns=COLUMN_ORDER)
        if category != "bancor-pero-no-enviamos" and "MOTIVO" in df.columns:
            df = df.drop(columns=["MOTIVO"])
        columns = [col for col in COLUMN_ORDER if col in df.columns]
        df = df.reindex(columns=columns)
        path = next_available_path(output_dir / EXPORT_FILENAMES[category])
        export_to_excel(df, path)
        results[category] = ExportFileResult(path=path, row_count=len(df))
    return results


REPORT_DECIMAL_COLUMNS: Sequence[str] = (
    "Mes Actual - Monto Total",
    "Mes Actual - Disparo Máximo",
    "Mes Pasado - Total Enviado",
    "Mes Pasado - Máximo Cobrado",
    "Mes Pasado - Total Cobrado",
    "Mes Pasado - Total sin Cobrar",
    "Monto Vencido Mas Última Cuota",
    "Valor Cuota",
    "Cuota y Media",
)


def write_report_file(
    report_rows: Sequence[Dict[str, object]],
    output_dir: Path,
) -> ExportFileResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(report_rows) if report_rows else pd.DataFrame(columns=REPORT_COLUMNS)
    df = df.reindex(columns=REPORT_COLUMNS)
    for column in REPORT_DECIMAL_COLUMNS:
        if column in df.columns:
            df[column] = df[column].apply(lambda value: float(_to_decimal(value)))
    path = next_available_path(output_dir / REPORT_FILENAME)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Reporte", index=False)
        worksheet = writer.sheets["Reporte"]
        row_count = len(df)
        if row_count:
            worksheet.insert_rows(1)
            column_lookup = {name: idx + 1 for idx, name in enumerate(df.columns)}
            data_start_row = 3
            data_end_row = row_count + 2

            monetary_columns = [
                "Mes Actual - Monto Total",
                "Mes Actual - Disparo Máximo",
                "Mes Pasado - Total Enviado",
                "Mes Pasado - Máximo Cobrado",
                "Monto Vencido Mas Última Cuota",
                "Valor Cuota",
                "Cuota y Media",
            ]
            count_columns = [
                "Planillas por Socio",
                "Mes Actual - Cantidad Disparos",
                "Mes Pasado - Cantidad Disparos",
            ]

            def iter_rows(col_name: str) -> list[str]:
                col_idx = column_lookup[col_name]
                col_letter = get_column_letter(col_idx)
                return [f"{col_letter}{row}" for row in range(data_start_row, data_end_row + 1)]

            for col_name in monetary_columns:
                if col_name in column_lookup:
                    for cell_ref in iter_rows(col_name):
                        worksheet[cell_ref].number_format = "#,##0.00"

            for col_name in count_columns:
                if col_name in column_lookup:
                    for cell_ref in iter_rows(col_name):
                        worksheet[cell_ref].number_format = "0"

            current_block = [
                "Mes Actual - ¿Se Envía?",
                "Mes Actual - Monto Total",
                "Mes Actual - Disparo Máximo",
                "Mes Actual - Cantidad Disparos",
                "Mes Actual - Detalle Disparos",
            ]
            past_block = [
                "Mes Pasado - ¿Se Envió?",
                "Mes Pasado - Total Enviado",
                "Mes Pasado - Máximo Cobrado",
                "Mes Pasado - Cantidad Disparos",
                "Mes Pasado - Respuestas",
                "Mes Pasado - Detalle Disparos",
    "Mes Pasado - Total Cobrado",
    "Mes Pasado - Total sin Cobrar",
            ]

            current_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            past_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            header_font = Font(bold=True)

            column_span_current = [column_lookup[name] for name in current_block if name in column_lookup]
            column_span_past = [column_lookup[name] for name in past_block if name in column_lookup]
            for col_name in df.columns:
                col_idx = column_lookup[col_name]
                header_cell = worksheet.cell(row=2, column=col_idx)
                display_value = col_name
                if col_name.startswith("Mes Actual - "):
                    display_value = col_name.split("Mes Actual - ", 1)[1]
                elif col_name.startswith("Mes Pasado - "):
                    display_value = col_name.split("Mes Pasado - ", 1)[1]
                header_cell.value = display_value
                header_cell.font = header_font
                top_cell = worksheet.cell(row=1, column=col_idx)
                if not isinstance(top_cell, MergedCell):
                    top_cell.value = display_value
                    top_cell.font = header_font

            if column_span_current:
                start_idx, end_idx = min(column_span_current), max(column_span_current)
                worksheet.merge_cells(start_row=1, start_column=start_idx, end_row=1, end_column=end_idx)
                top_cell = worksheet.cell(row=1, column=start_idx, value="Mes Actual")
                top_cell.font = header_font
                top_cell.fill = current_fill
                for idx in range(start_idx, end_idx + 1):
                    worksheet.cell(row=2, column=idx).fill = current_fill

            if column_span_past:
                start_idx, end_idx = min(column_span_past), max(column_span_past)
                worksheet.merge_cells(start_row=1, start_column=start_idx, end_row=1, end_column=end_idx)
                top_cell = worksheet.cell(row=1, column=start_idx, value="Mes Pasado")
                top_cell.font = header_font
                top_cell.fill = past_fill
                for idx in range(start_idx, end_idx + 1):
                    worksheet.cell(row=2, column=idx).fill = past_fill

            grouped_columns = set(current_block + past_block)
            for col_name in df.columns:
                if col_name in grouped_columns:
                    continue
                col_idx = column_lookup[col_name]
                header_value = worksheet.cell(row=2, column=col_idx).value
                top_cell = worksheet.cell(row=1, column=col_idx)
                if not isinstance(top_cell, MergedCell):
                    top_cell.value = header_value
                    top_cell.font = header_font

            def apply_conditional(column_name: str, true_color: str, false_color: str) -> None:
                if column_name not in column_lookup:
                    return
                col_idx = column_lookup[column_name]
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
                worksheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="equal",
                        formula=['"SI"'],
                        fill=PatternFill(start_color=true_color, end_color=true_color, fill_type="solid"),
                    ),
                )
                worksheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="equal",
                        formula=['"NO"'],
                        fill=PatternFill(start_color=false_color, end_color=false_color, fill_type="solid"),
                    ),
                )


            def apply_value_fill(column_name: str, color_map: Dict[str, str]) -> None:
                if column_name not in column_lookup:
                    return
                col_idx = column_lookup[column_name]
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
                for value, color in color_map.items():
                    worksheet.conditional_formatting.add(
                        cell_range,
                        CellIsRule(
                            operator="equal",
                            formula=[f'"{value}"'],
                            fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                        ),
                    )

            apply_conditional("Mes Actual - ¿Se Envía?", "C6EFCE", "FFC7CE")
            apply_conditional("Mes Pasado - ¿Se Envió?", "C6EFCE", "FFC7CE")

            apply_value_fill(
                "Clasificación CBU",
                {
                    "Bancor": "C6EFCE",
                    "No Bancor": "FFC7CE",
                    "Nacion": "C6EFCE",
                    "No Nacion": "FFC7CE",
                    "A revisar": "FFF2CC",
                },
            )
            apply_value_fill(
                "Clasificación Caja40",
                {"GRUPOS 1-20": "BDD7EE", "ARRASTRE": "DDEBF7", "CAJAS SIN USO": "FFC7CE"},
            )

            last_column_letter = get_column_letter(len(df.columns))
            worksheet.auto_filter.ref = f"A2:{last_column_letter}{data_end_row}"
            worksheet.freeze_panes = "B3"

    return ExportFileResult(path=path, row_count=len(df))


def log_discarded_entries(entries: List[Dict[str, object]], output_dir: Path) -> Optional[Path]:
    if not entries:
        return None
    log_path = output_dir / DISCARDED_LOG_NAME
    with log_path.open("w", encoding="utf-8") as handle:
        for entry in entries:
            handle.write(json.dumps(entry, ensure_ascii=False) + "\n")
    logger.info("Registradas %s planillas descartadas en %s", len(entries), log_path)
    return log_path


def generate_report(
    *,
    report_date: Optional[_dt.date] = None,
    filter_expr: Optional[str] = None,
    fields: Sequence[str] = FIELDS,
    max_rows: int = MAX_ROWS,
    input_excel: Path | str = "example-input.xlsx",
    output_dir: Path | str = "output",
    dev_mode: bool = False,
    arrastre_mode: bool = False,
    club_mutual_mode: bool = False,
    progress_callback: Optional[ProgressCallback] = None,
    api_dump_path: Path | None = None,
) -> ExportResult:
    def notify(progress: float, message: str) -> None:
        if progress_callback:
            progress_callback(max(0.0, min(progress, 1.0)), message)

    input_excel_path = Path(input_excel)
    output_dir_path = Path(output_dir)
    output_dir_path.mkdir(parents=True, exist_ok=True)
    if report_date is None:
        report_date = month_end_for(_dt.date.today())
    else:
        report_date = month_end_for(report_date)

    effective_filter = filter_expr or build_filter_for_date(
        report_date,
        line_superior=linea_superior_for_mode(club_mutual_mode=club_mutual_mode),
    )

    notify(0.05, "Validando insumos...")
    attempts = read_excel_attempts(input_excel_path)

    notify(0.2, "Solicitando datos del API...")
    if api_dump_path is not None:
        rows, raw_text = load_rows_from_dump(Path(api_dump_path))
    else:
        fetch_result = fetch_rows(
            filter_expr=effective_filter,
            fields=fields,
            max_rows=max_rows,
            return_raw_text=True,
        )
        rows, raw_text = fetch_result
        if not rows:
            raise RuntimeError("EvaluateList returned no rows for the given criteria.")
    if not rows:
        raise RuntimeError("EvaluateList returned no rows for the given criteria.")

    raw_response_path = next_available_path(
        output_dir_path / f"evaluate_list_response_{report_date:%Y%m%d}.txt"
    )
    raw_response_path.write_text(raw_text, encoding="utf-8")

    rows = sort_api_rows(rows)

    notify(0.4, f"Procesando {len(rows)} cuotas recibidas...")
    api_rows = parse_api_rows(rows)
    api_rows = filter_api_rows_for_mode(api_rows, club_mutual_mode=club_mutual_mode)
    if not api_rows:
        if club_mutual_mode:
            raise RuntimeError(
                "EvaluateList no devolvio cuotas para la linea superior "
                f"{CLUB_MUTUAL_LINEA_SUPERIOR}."
            )
        raise RuntimeError("EvaluateList returned no parsable rows for the given criteria.")
    planillas = consolidate_planillas(api_rows, attempts)

    notify(0.6, f"Clasificando {len(planillas)} planillas...")
    outputs, discarded_entries, report_rows = classify_planillas(
        planillas,
        report_date,
        dev_mode=dev_mode,
        arrastre_mode=arrastre_mode,
        club_mutual_mode=club_mutual_mode,
    )

    notify(0.8, "Generando archivos Excel...")
    file_results = write_output_files(outputs, output_dir_path)
    report_file_result = write_report_file(report_rows, output_dir_path)
    file_results["reporte"] = report_file_result

    discarded_log_path = log_discarded_entries(discarded_entries, output_dir_path) if dev_mode else None

    notify(1.0, "Exportacion completada.")
    return ExportResult(
        files=file_results,
        discarded_entries=discarded_entries,
        discarded_log=discarded_log_path,
        api_response_path=raw_response_path,
    )
